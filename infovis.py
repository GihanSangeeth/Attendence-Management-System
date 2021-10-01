import openpyxl
import re
from openpyxl import load_workbook
import matplotlib.pyplot as plt


def drawGraf(countP,countAb):
    left = [1, 2]
    height = [countP,countAb]
    tick_label = ['Present', 'Absent']
    plt.bar(left, height, tick_label = tick_label,width = 0.8, color = ['red', 'green'])
    plt.xlabel('x - axis')
    plt.ylabel('y - axis')
    plt.title('Attendance chart!')
    plt.show()

def grafChart(countP,countAb):
    summary = [countP,countAb]
    grade = ['Present', 'Absent']
    cl = ['green','red']
    plt.pie(summary, labels=grade, autopct='%2.1f%%', colors=cl)
    plt.show()
    
    
path = 'attendance.xlsx'
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

arr=[]
for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 2)
    a_string = cell_obj.value
    
    for word in a_string.split():
        if word.isdigit():
            number=int(word)   
            arr.append(number)
            print(number)

stId = input('enter student Id:')

count = 2
found = ''
for x in arr:
    if str(x)==stId:
        found = 'found'
        break    
    count+=1

    
if(found == 'found'):
    ws = wb_obj.worksheets[0]
    countP = 0
    countAb = 0
    for row in ws.iter_rows(min_row=count, max_row=count, values_only=True):
        for i in row:
            if i == 'Present':
                countP+=1
            elif i == 'Absent':
                countAb+=1    
    drawGraf(countP,countAb)
    grafChart(countP,countAb)
#     print(countP,countAb)
else:
    print('Not Found')




    

