import cv2
import numpy as np
import pytesseract
import pandas as pd
import matplotlib.pyplot as plt
import imutils
import xml.etree.cElementTree as ET
from PIL import Image
from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter
from openpyxl.utils import get_column_letter
import os.path
import glob

pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'





#------- create image threshold-------------------------------
def preProcessing(img):
    imgGray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    imgBlur = cv2.GaussianBlur(imgGray,(5,5),6)
    imgThres = cv2.adaptiveThreshold(imgBlur,255,1,1,11,2)
    
    return imgThres


#----- find biggest contour----------------------------------
def biggestContour(contours):
    biggest = np.array([])
    maxArea = 0
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area>60:
            peri = cv2.arcLength(cnt,True)
            approx = cv2.approxPolyDP(cnt,0.02*peri,True)
            if area > maxArea and len(approx) == 4:
                biggest = approx
                maxArea = area
    return biggest,maxArea

#-----reorder --------------------------------------------
def reorder(myPoints):
    myPoints = myPoints.reshape((4,2))
    myPointsNew = np.zeros((4,1,2), dtype=np.int32)
    add = myPoints.sum(1)

    
    myPointsNew[0] = myPoints[np.argmin(add)]
    myPointsNew[3] = myPoints[np.argmax(add)]
    
    diff = np.diff(myPoints, axis = 1)
    myPointsNew[1] = myPoints[np.argmin(diff)]
    myPointsNew[2] = myPoints[np.argmax(diff)]
    
    return myPointsNew


# -------Short contours----------------------------
def sort_contours(cnts, method="left-to-right"):
	reverse = False
	i = 0
	if method == "right-to-left" or method == "bottom-to-top":
		reverse = True
	if method == "top-to-bottom" or method == "bottom-to-top":
		i = 1
	boundingBoxes = [cv2.boundingRect(c) for c in cnts]
	(cnts, boundingBoxes) = zip(*sorted(zip(cnts, boundingBoxes),
		key=lambda b:b[1][i], reverse=reverse))
	return (cnts, boundingBoxes)


# ---------Find Vertical and Horizontal lines in table-------------------------

def findLines(image,vl,vi,hl,hi,x,y,z):
    read_image = cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
    convert_bin,grey_scale = cv2.threshold(read_image,128,255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    grey_scale = 255-grey_scale
#     grey_graph = plt.imshow(grey_scale,cmap='gray')
    length = np.array(read_image).shape[1]//z
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (length, hl))
    horizontal_detect = cv2.erode(grey_scale, horizontal_kernel, iterations=hi)
    hor_lines = cv2.dilate(horizontal_detect, horizontal_kernel, iterations=hi)

    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (vl, length))
    vertical_detect = cv2.erode(grey_scale, vertical_kernel, iterations=vi)
    ver_lines = cv2.dilate(vertical_detect, vertical_kernel, iterations=vi)
    
    alpha = 0.5
    beta = 1.0 - alpha
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (x, y))
    img_final_bin = cv2.addWeighted(ver_lines, alpha, hor_lines, beta, 0.0)
    img_final_bin = cv2.erode(~img_final_bin, kernel, iterations=3)
    (thresh, img_final_bin) = cv2.threshold(img_final_bin, 128,255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
    contoursT, hierarchyT = cv2.findContours(img_final_bin, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    
    return contoursT,img_final_bin


#------------ Student Attendance--------------
def findAttendence(img):

    signImgGray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    signImgTresh = cv2.threshold(signImgGray,178,255,cv2.THRESH_BINARY_INV)[1]
    
    totalPixels = cv2.countNonZero(signImgTresh)
    print(totalPixels)
    if totalPixels > 2000 :
        status = 'Present'
    else:
        status = 'Absent'
    
    return status


# ----------create excel file -----------------------------
def createExcel(names,ids,att):
    if os.path.isfile('attendance.xlsx'):
        print ("File exist")
        filename = 'attendance.xlsx'
        wb=load_workbook(filename)
        ws = wb.worksheets[0]
        maxCol = ws.max_column
        nextLetter = get_column_letter(maxCol+1)
#         date = input("enter date:")
        ws[nextLetter+'1'] = 'Day'+str(day)
        x = 2
        for i in range(len(att)): 
            ws[nextLetter+str(x)] = att[i]
            x+=1
        wb.save(filename)
    else:
        y = 1
        filename = 'attendance.xlsx'
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({'bold': 1})
        worksheet.set_column(1, 10, 20)
        worksheet.write('A1', 'No', bold)
        worksheet.write('B1', 'Student Id', bold)
        worksheet.write('C1', 'Student Name', bold)
    
        for x in range(len(names)):    
            worksheet.write_number(y, 0 , y )
            worksheet.write_string(y, 1 , ids[x] ) 
            worksheet.write_string(y, 2 , names[x] )
            y +=1
        workbook.close()      
        
        wb=load_workbook(filename)
        ws = wb.worksheets[0]
        maxCol = ws.max_column
        nextLetter = get_column_letter(maxCol+1)
        ws[nextLetter+'1'] = 'Day'+str(day)
        x = 2
        for i in range(len(att)): 
            ws[nextLetter+str(x)] = att[i]
            x+=1
        wb.save(filename)

        
# ---------- create xml file with students data -----------------
def createXML(names,ids):
    
    root = ET.Element("University")
    
    no= 1
    for y in range(len(names)):  
        doc = ET.SubElement(root, "Students")
        ET.SubElement(doc, "NO", name="NO").text = str(no)
        ET.SubElement(doc, "StudentID", name="StudentID").text = ids[y]
        ET.SubElement(doc, "StudentName", name="StudentName").text = names[y]
        no+=1
    tree = ET.ElementTree(root)
    tree.write("info.xml")
    
#--------save signatures of stunets ------------
def saveSignatures(signature,no):
    folder = 'testImage'
    count = 1
    for z in signature:
        path = str(count)
        imgName = str(no)+ '.png'
        joined_path = os.path.join(folder,path,imgName)
        cv2.imwrite(joined_path, z)
        count+=1
    
    
    
path = glob.glob('assets/*.jpeg')
day = 1
for image in path:
    img = cv2.imread(image)
    h,w,c = img.shape
    widthImg = w//4
    heightImg = h//4
    img = cv2.resize(img,(widthImg,heightImg))
    
    
    imgBlank = np.zeros((heightImg, widthImg, 3), np.uint8)
    imgThres = preProcessing(img)
    
#     cv2.imshow('Threshold Image'+str(day),imgThres)
    

    imgContour = img.copy()
    imgBigContour = img.copy()
    contours, hierarchy = cv2.findContours(imgThres,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_NONE)
    cv2.drawContours(imgContour,contours,-1,(255,0,0),3)
    
#     cv2.imshow('Countours'+str(day),imgContour)
    
    biggest, maxArea = biggestContour(contours)


    


    names=[]
    ids=[]
    att = []
    signature = []
    if biggest.size != 0:
        biggest = reorder(biggest)
        cv2.drawContours(imgBigContour,biggest,-1,(0,0,255),20)
        pts1 = np.float32(biggest)
        pts2 = np.float32([[0,0],[widthImg,0],[0,heightImg],[widthImg,heightImg]])
        matrix= cv2.getPerspectiveTransform(pts1, pts2)
        imgWarpColored = cv2.warpPerspective(img,matrix,(widthImg,heightImg))
    
        imageNew = imgWarpColored.copy()
        imageNew = cv2.resize(imageNew,(720,240))
#         cv2.imshow("Cropped Table",imageNew)
    
        contoursT,img_final_bin1=findLines(imageNew,vl=10,vi=10,hl=1,hi=1,x=3,y=1,z=10)
#         cv2.imshow("Final Lines",img_final_bin1)
        
        (contoursnew, boundingBoxes) = sort_contours(contoursT, method="top-to-bottom")
        
        
        idx = 0
        images= []
        for c in contoursnew:
            x, y, w, h = cv2.boundingRect(c)
            if (w > 100 and h > 20 ) and w > 3*h:
                idx +=1
                new_img = imageNew[y:y+h, x:x+w]
                images.append(new_img)
#                 cv2.imshow("img"+str(idx),new_img)
        no = 1      
        for i in range(len(images)):
            idx += 1
            if i ==0:pass
            else:
                col = images[i].copy()
                contoursSt,img_final_bin2=findLines(col,vl=1,vi=10,hl=10,hi=20,x=2,y=2,z=100)
#                 cv2.imshow("Final Lines",img_final_bin2)
                (contoursS, boundingBoxesS) = sort_contours(contoursSt, method="left-to-right")
                idx2 = 0
                cells= []
                myData=[]
                for c in contoursS:
                    x, y, w, h = cv2.boundingRect(c)
                    if (w > 40 and h > 10):
                        idx2 += 1
                        #new_img2 = col[y+3:y+h-3, x:x+w-3] 
                        new_img2 = col[y:y+h-3, x:x+w-2]
                        new_img2 = cv2.resize(new_img2, None, fx=3, fy=3, interpolation=cv2.
                                              INTER_CUBIC)
                        grayT = cv2.cvtColor(new_img2, cv2.COLOR_BGR2GRAY)
                        grayT = cv2.bitwise_not(grayT)
                        threshT = cv2.threshold(grayT, 0, 255,cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
                        cells.append(new_img2)
                        myData.append(pytesseract.image_to_string(threshT))
                

                names.append(myData[3])
                ids.append(myData[1])
                status = findAttendence(cells[4])
                signature.append(cells[4])
                att.append(status)             
                no +=1
    
    
    createXML(names,ids)
    createExcel(names,ids,att)
    saveSignatures(signature,day)

    day +=1
    cv2.waitKey(0)
print('End')
    

